import re
import os
import pandas as pd

from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from datetime import datetime
from openpyxl import load_workbook

# =========================================================
# CONFIG
# =========================================================

URL = "https://www.tenable.com/security/research"

OUTPUT_FILE = "tenable_cves.xlsx"

CVE_PATTERN = r"CVE-\d{4}-\d{4,7}"

# =========================================================
# LOAD EXISTING KEYS
# Prevent SAME ROW duplicate
# But allow same CVE on different dates
# =========================================================

def load_existing_keys():

    if not os.path.exists(OUTPUT_FILE):
        return set()

    wb = load_workbook(OUTPUT_FILE)

    ws = wb.active

    keys = set()

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        try:

            key = (
                str(row[0]),  # CVE
                str(row[1]),  # DATE
                str(row[2])   # ADVISORY
            )

            keys.add(key)

        except:
            continue

    return keys

# =========================================================
# LOAD EXISTING DATAFRAME
# =========================================================

def load_existing_df():

    if not os.path.exists(OUTPUT_FILE):

        return pd.DataFrame(
            columns=[
                "CVE",
                "DATE",
                "ADVISORY_ID",
                "NAME",
                "SEVERITY",
                "SCRAPED_AT"
            ]
        )

    return pd.read_excel(OUTPUT_FILE)

# =========================================================
# MAIN
# =========================================================

def main():

    print("\nOpening Tenable Research Page...\n")

    existing_keys = load_existing_keys()

    existing_df = load_existing_df()

    new_rows = []

    scraped_at = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=True
        )

        page = browser.new_page()

        page.goto(
            URL,
            wait_until="domcontentloaded",
            timeout=60000
        )

        # =====================================================
        # SCROLL FULL PAGE
        # =====================================================

        previous_height = 0

        while True:

            page.evaluate(
                "window.scrollTo(0, document.body.scrollHeight)"
            )

            page.wait_for_timeout(3000)

            current_height = page.evaluate(
                "document.body.scrollHeight"
            )

            if current_height == previous_height:
                break

            previous_height = current_height

        print("\nFull page loaded successfully")

        html = page.content()

        browser.close()

    # =========================================================
    # PARSE HTML
    # =========================================================

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    tr_list = soup.find_all("tr")

    print("\nRows Found:", len(tr_list))

    for tr in tr_list:

        tds = tr.find_all("td")

        if len(tds) < 5:
            continue

        try:

            # =================================================
            # DATE
            # =================================================

            date_text = tds[0].get_text(
                " ",
                strip=True
            )

            # =================================================
            # ADVISORY ID
            # =================================================

            advisory_id = tds[1].get_text(
                " ",
                strip=True
            )

            # =================================================
            # NAME
            # =================================================

            name = tds[2].get_text(
                " ",
                strip=True
            )

            # =================================================
            # SEVERITY
            # =================================================

            severity = tds[3].get_text(
                " ",
                strip=True
            )

            # =================================================
            # CVE TEXT
            # =================================================

            cve_text = tds[4].get_text(
                " ",
                strip=True
            )

            # =================================================
            # EXTRACT ALL CVEs
            # =================================================

            found_cves = re.findall(
                CVE_PATTERN,
                cve_text,
                re.IGNORECASE
            )

            if not found_cves:
                continue

            # =================================================
            # ONE CVE = ONE ROW
            # =================================================

            for cve in found_cves:

                cve = cve.upper().strip()

                key = (
                    cve,
                    date_text,
                    advisory_id
                )

                # =============================================
                # SKIP EXACT DUPLICATES
                # =============================================

                if key in existing_keys:
                    continue

                existing_keys.add(key)

                print("\n" + "=" * 60)
                print("DATE      :", date_text)
                print("ADVISORY  :", advisory_id)
                print("NAME      :", name)
                print("SEVERITY  :", severity)
                print("CVE       :", cve)

                new_rows.append({
                    "CVE": cve,
                    "DATE": date_text,
                    "ADVISORY_ID": advisory_id,
                    "NAME": name,
                    "SEVERITY": severity,
                    "SCRAPED_AT": scraped_at
                })

        except Exception as e:

            print("\nERROR PARSING ROW")
            print(e)

    # =========================================================
    # APPEND NEW DATA
    # =========================================================

    if new_rows:

        new_df = pd.DataFrame(new_rows)

        final_df = pd.concat(
            [existing_df, new_df],
            ignore_index=True
        )

        final_df.to_excel(
            OUTPUT_FILE,
            index=False
        )

        print("\n" + "=" * 60)
        print("NEW ROWS ADDED :", len(new_rows))
        print("TOTAL ROWS     :", len(final_df))
        print("EXCEL UPDATED  :", OUTPUT_FILE)
        print("=" * 60)

    else:

        print("\nNo new rows found")

# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":
    main()
